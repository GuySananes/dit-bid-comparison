"""Layer 1 — Excel parser.

Converts any contractor Excel file into a ParsedFile with zero AI calls.
Handles: flexible header detection, merged cells, keyword-based column mapping,
section-header skipping, and flag detection.
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from src.layer1_parser.schema import FileMeta, ParsedFile, ParsedRow, ParsedSheet, RowFlags

# ---------------------------------------------------------------------------
# Column keyword maps — order matters: first match wins
# ---------------------------------------------------------------------------
DESCRIPTION_KEYWORDS  = ["תאור", "תיאור", "description"]
UNIT_KEYWORDS         = ["יחידה", "unit"]
QUANTITY_KEYWORDS     = ["כמות", "qty", "quantity"]
UNIT_PRICE_KEYWORDS   = ["מחיר יחידה", "unit price"]
TOTAL_PRICE_KEYWORDS  = ['סה"כ', "total"]
MANUFACTURER_KEYWORDS = ["יצרן", "manufacturer", "model"]
MKT_KEYWORDS          = ['מק"ט', "mkt", "part number", "הערות"]

COLUMN_MAP = {
    "description":        DESCRIPTION_KEYWORDS,
    "unit":               UNIT_KEYWORDS,
    "quantity":           QUANTITY_KEYWORDS,
    "unit_price":         UNIT_PRICE_KEYWORDS,
    "total_price":        TOTAL_PRICE_KEYWORDS,
    "manufacturer_model": MANUFACTURER_KEYWORDS,
    "mkt_raw":            MKT_KEYWORDS,
}

# Flag trigger strings
_EXISTING_MARKERS = ["ציוד קיים", "קיים"]
_NOT_IN_TOTAL_MARKERS = ["לא לסיכום", "אופציה"]
_OPTIONAL_MARKERS = ["אופציה"]

# Sheet-level skip / override triggers
_SUMMARY_SHEET_NAMES = ["סיכום"]
_EXISTING_EQUIPMENT_SHEET_MARKER = "ציוד קיים"

MATH_TOLERANCE = 0.01  # 1 %


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _cell_str(value) -> str:
    """Return cell value as stripped string, or empty string."""
    if value is None:
        return ""
    return str(value).strip()


def _contains_any(text: str, markers: list[str]) -> bool:
    text_lower = text.lower()
    return any(m.lower() in text_lower for m in markers)


def _to_float(value) -> Optional[float]:
    """Try to convert a cell value to float; return None on failure."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value).replace(",", "."))
    try:
        return float(cleaned)
    except ValueError:
        return None


def _unmerge_sheet(ws) -> None:
    """Expand merged cells so every cell in the range holds the top-left value."""
    for merged_range in list(ws.merged_cells.ranges):
        top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in ws.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                cell.value = top_left


def _detect_header_row(ws, max_scan: int = 30) -> Optional[int]:
    """Return the 1-based row index of the header row, or None if not found."""
    all_keywords = [kw for kws in COLUMN_MAP.values() for kw in kws]
    best_row, best_count = None, 0

    for row_idx in range(1, min(max_scan, ws.max_row) + 1):
        row_text = " ".join(
            _cell_str(ws.cell(row_idx, col).value).lower()
            for col in range(1, ws.max_column + 1)
        )
        count = sum(1 for kw in all_keywords if kw.lower() in row_text)
        if count > best_count:
            best_count = count
            best_row = row_idx

    # Require at least 2 recognised keywords to consider it a real header
    return best_row if best_count >= 2 else None


def _map_columns(ws, header_row: int) -> dict[str, Optional[int]]:
    """Return {field_name: col_index (1-based)} for the detected header row."""
    mapping: dict[str, Optional[int]] = {field: None for field in COLUMN_MAP}

    for col in range(1, ws.max_column + 1):
        cell_text = _cell_str(ws.cell(header_row, col).value).lower()
        for field, keywords in COLUMN_MAP.items():
            if mapping[field] is not None:
                continue  # already assigned
            if any(kw.lower() in cell_text for kw in keywords):
                mapping[field] = col
                break

    return mapping


def _is_section_header(row_values: list, col_map: dict) -> bool:
    """True when a row looks like a section label with no numeric price data."""
    # A section header typically has a description but no numeric total_price
    total_col = col_map.get("total_price")
    unit_price_col = col_map.get("unit_price")
    qty_col = col_map.get("quantity")

    def _val(col):
        if col is None or col > len(row_values):
            return None
        return row_values[col - 1]

    has_total = _to_float(_val(total_col)) is not None
    has_unit_price = _to_float(_val(unit_price_col)) is not None
    has_qty = _to_float(_val(qty_col)) is not None

    return not (has_total or has_unit_price or has_qty)


def _detect_flags(
    description: str,
    unit_price_raw,
    total_price_raw,
    notes: str,
    quantity: float,
    unit_price_float: Optional[float],
    total_price_float: Optional[float],
) -> RowFlags:
    up_str = _cell_str(unit_price_raw)
    tp_str = _cell_str(total_price_raw)

    existing = _contains_any(up_str, _EXISTING_MARKERS) or _contains_any(tp_str, _EXISTING_MARKERS)
    not_in_total = _contains_any(tp_str, _NOT_IN_TOTAL_MARKERS)
    optional = _contains_any(description, _OPTIONAL_MARKERS) or _contains_any(notes, _OPTIONAL_MARKERS)

    math_error = False
    if (
        unit_price_float is not None
        and total_price_float is not None
        and total_price_float != 0
        and quantity > 0
    ):
        expected = quantity * unit_price_float
        diff = abs(expected - total_price_float) / abs(total_price_float)
        math_error = diff > MATH_TOLERANCE

    return RowFlags(
        existing_equipment=existing,
        not_in_total=not_in_total,
        optional=optional,
        math_error=math_error,
    )


def _is_summary_sheet(ws, sheet_name: str) -> bool:
    """True for sheets that are aggregation-only and should be skipped."""
    # Name-based check
    if any(name in sheet_name for name in _SUMMARY_SHEET_NAMES):
        return True
    # Structure-based check: no unit/price columns found after scanning
    _unmerge_sheet(ws)
    header_row = _detect_header_row(ws)
    if header_row is None:
        return True
    col_map = _map_columns(ws, header_row)
    has_price_col = col_map.get("unit_price") is not None or col_map.get("total_price") is not None
    return not has_price_col


def _parse_sheet(ws, sheet_name: str) -> Optional[ParsedSheet]:
    """Parse one worksheet into a ParsedSheet, or None if it should be skipped."""
    # _is_summary_sheet calls _unmerge_sheet internally, so we skip calling it again
    if _is_summary_sheet(ws, sheet_name):
        return None

    # Merged cells already expanded by _is_summary_sheet — do not call again.
    # All rows in a sheet named "ציוד קיים" are existing equipment by definition.
    sheet_is_existing = _EXISTING_EQUIPMENT_SHEET_MARKER in sheet_name

    header_row = _detect_header_row(ws)
    if header_row is None:
        return None

    col_map = _map_columns(ws, header_row)

    def _get(row_values: list, field: str):
        col = col_map.get(field)
        if col is None or col > len(row_values):
            return None
        return row_values[col - 1]

    rows: list[ParsedRow] = []
    data_row_index = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]

        # Skip entirely blank rows
        if all(v is None for v in row_values):
            continue

        # Skip section headers (no numeric price data)
        if _is_section_header(row_values, col_map):
            continue

        # Raw values
        desc_raw       = _get(row_values, "description")
        unit_raw       = _get(row_values, "unit")
        qty_raw        = _get(row_values, "quantity")
        up_raw         = _get(row_values, "unit_price")
        tp_raw         = _get(row_values, "total_price")
        mfr_raw        = _get(row_values, "manufacturer_model")
        mkt_raw        = _get(row_values, "mkt_raw")

        description = _cell_str(desc_raw)
        if not description:
            continue  # nothing meaningful on this row

        unit           = _cell_str(unit_raw)
        quantity       = _to_float(qty_raw) or 0.0
        up_float       = _to_float(up_raw)
        tp_float       = _to_float(tp_raw)

        # unit_price: 0.0 when empty, raw string when non-numeric text (e.g. "ציוד קיים")
        up_str_raw = _cell_str(up_raw)
        if up_float is not None:
            unit_price: float | str = up_float
        elif up_str_raw == "":
            unit_price = 0.0
        else:
            unit_price = up_str_raw

        # total_price: keep raw string when non-numeric (e.g. "ציוד קיים", "לא לסיכום")
        total_price: float | str = tp_float if tp_float is not None else _cell_str(tp_raw)

        # notes: any remaining non-empty cells not mapped to a known column
        mapped_cols = {c for c in col_map.values() if c is not None}
        notes_parts = [
            _cell_str(row_values[c - 1])
            for c in range(1, ws.max_column + 1)
            if c not in mapped_cols and _cell_str(row_values[c - 1])
        ]
        notes = "; ".join(notes_parts)

        flags = _detect_flags(
            description=description,
            unit_price_raw=up_raw,
            total_price_raw=tp_raw,
            notes=notes,
            quantity=quantity,
            unit_price_float=up_float,
            total_price_float=tp_float,
        )
        if sheet_is_existing:
            flags = flags.model_copy(update={"existing_equipment": True})

        data_row_index += 1
        rows.append(
            ParsedRow(
                row_index=data_row_index,
                description=description,
                unit=unit,
                quantity=quantity,
                unit_price=unit_price,
                total_price=total_price,
                manufacturer_model=_cell_str(mfr_raw),
                mkt_raw=_cell_str(mkt_raw),
                notes=notes,
                flags=flags,
            )
        )

    sheet_total = sum(
        r.total_price for r in rows if isinstance(r.total_price, float)
    )

    return ParsedSheet(sheet_name=sheet_name, rows=rows, sheet_total=sheet_total)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_excel(
    file_path: str | Path,
    contractor_id: str,
    project_id: str,
) -> ParsedFile:
    """Parse an Excel file and return a ParsedFile.

    Args:
        file_path: Path to the .xlsx contractor proposal.
        contractor_id: Short identifier for the contractor (e.g. "contractor_a").
        project_id: Project identifier (e.g. "proj_2026_001").

    Returns:
        ParsedFile with all sheets parsed.
    """
    file_path = Path(file_path)

    wb = openpyxl.load_workbook(str(file_path), data_only=True)

    sheets: list[ParsedSheet] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parsed = _parse_sheet(ws, sheet_name)
        if parsed is not None:
            sheets.append(parsed)

    meta = FileMeta(
        contractor_id=contractor_id,
        file_name=file_path.name,
        project_id=project_id,
        parsed_at=datetime.now(timezone.utc),
    )

    return ParsedFile(meta=meta, sheets=sheets)


def parse_excel_to_json(
    file_path: str | Path,
    contractor_id: str,
    project_id: str,
    output_path: Optional[str | Path] = None,
) -> str:
    """Parse an Excel file and return (and optionally save) JSON.

    Args:
        file_path: Path to the .xlsx contractor proposal.
        contractor_id: Short identifier for the contractor.
        project_id: Project identifier.
        output_path: If provided, write the JSON to this path.

    Returns:
        JSON string of the ParsedFile.
    """
    parsed = parse_excel(file_path, contractor_id, project_id)
    json_str = parsed.model_dump_json(indent=2)

    if output_path is not None:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json_str, encoding="utf-8")

    return json_str


# ---------------------------------------------------------------------------
# CLI convenience
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Parse a contractor Excel file to JSON.")
    parser.add_argument("file", help="Path to the .xlsx file")
    parser.add_argument("--contractor", required=True, help="Contractor ID")
    parser.add_argument("--project", required=True, help="Project ID")
    parser.add_argument("--output", help="Output JSON file path (prints to stdout if omitted)")
    args = parser.parse_args()

    result = parse_excel_to_json(args.file, args.contractor, args.project, args.output)
    if not args.output:
        print(result)
