"""Tests for Layer 1 — Excel parser.

Covers: header detection, merged cell handling, flag detection, math errors.

Synthetic workbooks are created with openpyxl so tests run on any machine.
Smoke tests against the real sample files are skipped when those files are absent.
"""

from pathlib import Path

import openpyxl
import pytest

from src.layer1_parser.excel_reader import parse_excel
from src.layer1_parser.schema import ParsedFile

# ── sample file paths ─────────────────────────────────────────────────────────

_RAW = Path("data/raw")
_FILE_A = _RAW / "הצעה 1.xlsx"
_FILE_B = _RAW / "הצעה 2.xlsx"

requires_samples = pytest.mark.skipif(
    not (_FILE_A.exists() and _FILE_B.exists()),
    reason="Sample Excel files not present in data/raw/",
)

# Column order that the parser recognises via keyword matching
_HEADERS = ["מספר", "תיאור", "יחידה", "כמות", "מחיר יחידה", 'סה"כ', "יצרן + דגם", 'מק"ט']


# ── workbook builder helper ───────────────────────────────────────────────────

def _write_xlsx(
    tmp_path: Path,
    sheet_rows: list,
    sheet_name: str = "חדר ישיבות",
    header_offset: int = 0,     # blank rows to insert before the header
) -> Path:
    """Write a minimal valid xlsx with recognised headers and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for _ in range(header_offset):
        ws.append([""] * len(_HEADERS))
    ws.append(_HEADERS)
    for row in sheet_rows:
        ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


# ── smoke tests on real sample files ─────────────────────────────────────────

@requires_samples
def test_parse_file_a_returns_parsedfile():
    result = parse_excel(_FILE_A, "contractor_a", "p1")
    assert isinstance(result, ParsedFile)
    assert len(result.sheets) >= 1


@requires_samples
def test_parse_file_b_returns_parsedfile():
    result = parse_excel(_FILE_B, "contractor_b", "p1")
    assert isinstance(result, ParsedFile)
    assert len(result.sheets) >= 1


@requires_samples
def test_all_sheets_have_rows():
    for path, cid in [(_FILE_A, "a"), (_FILE_B, "b")]:
        result = parse_excel(path, cid, "p1")
        for sheet in result.sheets:
            assert len(sheet.rows) >= 1, f"Sheet {sheet.sheet_name!r} has no rows"


@requires_samples
def test_meta_populated_correctly():
    result = parse_excel(_FILE_A, "contractor_a", "proj_test")
    assert result.meta.contractor_id == "contractor_a"
    assert result.meta.project_id == "proj_test"
    assert result.meta.file_name == _FILE_A.name


# ── header detection ──────────────────────────────────────────────────────────

def test_header_on_first_row(tmp_path):
    path = _write_xlsx(tmp_path, [
        ["", "מסך 65 אינץ", "יח", 1, 5000, 5000, "Samsung", "QB65C"],
    ])
    result = parse_excel(path, "c1", "p1")
    assert len(result.sheets[0].rows) == 1


def test_header_on_third_row(tmp_path):
    """Two blank rows before the header must not prevent detection."""
    path = _write_xlsx(tmp_path, [
        ["", "מסך 86 אינץ", "יח", 2, 8000, 16000, "LG", ""],
    ], header_offset=2)
    result = parse_excel(path, "c1", "p1")
    assert len(result.sheets[0].rows) == 1


def test_no_recognisable_header_produces_no_rows(tmp_path):
    """A sheet with unrelated column names yields zero data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "random"
    ws.append(["col1", "col2", "col3"])
    ws.append([1, 2, 3])
    path = tmp_path / "noheader.xlsx"
    wb.save(str(path))

    result = parse_excel(path, "c1", "p1")
    total_rows = sum(len(s.rows) for s in result.sheets)
    assert total_rows == 0


# ── merged cell handling ──────────────────────────────────────────────────────

def test_merged_section_header_does_not_crash(tmp_path):
    """A merged-cell section label between data rows must not raise."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "חדר ישיבות"

    for col, h in enumerate(_HEADERS, 1):
        ws.cell(1, col, h)

    # Merged section header spanning all columns
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(_HEADERS))
    ws.cell(2, 1).value = "קטגוריה א"

    # Regular data row after the merged section
    ws.append(["", "מסך 65 אינץ", "יח", 1, 5000, 5000, "Samsung", "QB65C"])

    path = tmp_path / "merged_section.xlsx"
    wb.save(str(path))

    result = parse_excel(path, "c1", "p1")
    assert isinstance(result, ParsedFile)
    # Merged section row has no numeric price → treated as section header → skipped
    assert len(result.sheets[0].rows) == 1


def test_merged_cells_in_data_column_handled(tmp_path):
    """Merged cells within the description column must not raise."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "אולפן"

    for col, h in enumerate(_HEADERS, 1):
        ws.cell(1, col, h)

    # Description merged across two rows
    ws.merge_cells(start_row=2, end_row=3, start_column=2, end_column=2)
    ws.cell(2, 2).value = "מסך 75 אינץ"
    ws.cell(2, 4).value = 1
    ws.cell(2, 5).value = 7000.0
    ws.cell(2, 6).value = 7000.0

    path = tmp_path / "merged_desc.xlsx"
    wb.save(str(path))

    result = parse_excel(path, "c1", "p1")
    assert isinstance(result, ParsedFile)


# ── flag detection ────────────────────────────────────────────────────────────

def test_existing_equipment_flag(tmp_path):
    """Unit price "ציוד קיים" triggers existing_equipment flag."""
    path = _write_xlsx(tmp_path, [
        ["", "מצלמה", "יח", 1, "ציוד קיים", "ציוד קיים", "Sony", ""],
    ])
    result = parse_excel(path, "c1", "p1")
    assert result.sheets[0].rows[0].flags.existing_equipment is True


def test_not_in_total_flag(tmp_path):
    """total_price "לא לסיכום" triggers not_in_total flag."""
    path = _write_xlsx(tmp_path, [
        ["", "ציוד אופציונלי", "יח", 1, 2000, "לא לסיכום", "Brand", ""],
    ])
    result = parse_excel(path, "c1", "p1")
    assert result.sheets[0].rows[0].flags.not_in_total is True


def test_optional_flag_from_description(tmp_path):
    """Description containing "אופציה" triggers optional flag."""
    path = _write_xlsx(tmp_path, [
        ["", "מסך אופציה", "יח", 1, 2000, 2000, "Sony", ""],
    ])
    result = parse_excel(path, "c1", "p1")
    assert result.sheets[0].rows[0].flags.optional is True


def test_math_error_flag_on_wrong_total(tmp_path):
    """3 × 4500 = 13500, stated total 13000 → math_error = True."""
    path = _write_xlsx(tmp_path, [
        ["", "מסך 65", "יח", 3, 4500, 13000, "Samsung", ""],
    ])
    result = parse_excel(path, "c1", "p1")
    assert result.sheets[0].rows[0].flags.math_error is True


def test_no_math_error_on_correct_total(tmp_path):
    path = _write_xlsx(tmp_path, [
        ["", "מסך 65", "יח", 3, 4500, 13500, "Samsung", ""],
    ])
    result = parse_excel(path, "c1", "p1")
    assert result.sheets[0].rows[0].flags.math_error is False


def test_math_error_within_tolerance(tmp_path):
    """Rounding of 0.5% difference must not trigger math_error (tolerance is 1%)."""
    path = _write_xlsx(tmp_path, [
        ["", "מסך 65", "יח", 1, 1000, 1005, "Samsung", ""],
    ])
    result = parse_excel(path, "c1", "p1")
    assert result.sheets[0].rows[0].flags.math_error is False


# ── row filtering ─────────────────────────────────────────────────────────────

def test_section_header_rows_skipped(tmp_path):
    """Text-only rows (no price) between data rows must be excluded."""
    path = _write_xlsx(tmp_path, [
        ["", "-- קטגוריה ראשית --", "", "", "", "", "", ""],  # section header
        ["", "מסך 65", "יח", 1, 5000, 5000, "Samsung", ""],   # real data
    ])
    result = parse_excel(path, "c1", "p1")
    assert len(result.sheets[0].rows) == 1


def test_blank_rows_skipped(tmp_path):
    path = _write_xlsx(tmp_path, [
        ["", "", "", "", "", "", "", ""],
        ["", "מסך 65", "יח", 1, 5000, 5000, "Samsung", ""],
        ["", "", "", "", "", "", "", ""],
    ])
    result = parse_excel(path, "c1", "p1")
    assert len(result.sheets[0].rows) == 1


def test_summary_sheet_skipped(tmp_path):
    """Sheets named 'סיכום' must be excluded from output."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "סיכום"
    for col, h in enumerate(_HEADERS, 1):
        ws.cell(1, col, h)
    ws.append(["", 'סה"כ כולל', "", "", "", 100000, "", ""])
    path = tmp_path / "summary.xlsx"
    wb.save(str(path))

    result = parse_excel(path, "c1", "p1")
    assert len(result.sheets) == 0


def test_multiple_data_sheets_all_parsed(tmp_path):
    """Each worksheet with a valid header gets its own ParsedSheet."""
    wb = openpyxl.Workbook()
    for name in ["חדר ישיבות", "אולפן"]:
        ws = wb.create_sheet(name)
        for col, h in enumerate(_HEADERS, 1):
            ws.cell(1, col, h)
        ws.append(["", f"פריט ב{name}", "יח", 1, 1000, 1000, "Brand", ""])
    # Remove default empty sheet
    del wb["Sheet"]
    path = tmp_path / "multi.xlsx"
    wb.save(str(path))

    result = parse_excel(path, "c1", "p1")
    assert len(result.sheets) == 2
    names = {s.sheet_name for s in result.sheets}
    assert "חדר ישיבות" in names
    assert "אולפן" in names
